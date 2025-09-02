import asyncio
from typing import Dict, Any, List, Optional
from app.agents.base_agent import BaseAgent
from app.models.pydantic_models import ProcessingResult
import logging
from pathlib import Path
import PyPDF2
import openpyxl
from PIL import Image
import pytesseract
import json
import re

logger = logging.getLogger(__name__)

class ExtractionAgent(BaseAgent):
    """Agent responsible for extracting information from various document types"""
    
    def __init__(self):
        super().__init__()
        self.supported_formats = ['.pdf', '.png', '.jpg', '.jpeg', '.xlsx', '.xls', '.txt']
        
    def get_capabilities(self) -> List[str]:
        return [
            "pdf_text_extraction",
            "image_ocr",
            "excel_data_extraction", 
            "text_parsing",
            "structured_data_extraction"
        ]
    
    async def process(self, input_data: Dict[str, Any]) -> ProcessingResult:
        """Process documents and extract structured information"""
        try:
            if not self.validate_input(input_data):
                return self.create_error_result("Invalid input data")
            
            documents = input_data.get('documents', [])
            if not documents:
                return self.create_error_result("No documents provided")
            
            extracted_data = []
            total_confidence = 0.0
            
            for doc in documents:
                doc_result = await self._process_document(doc)
                if doc_result.success:
                    extracted_data.append(doc_result.data)
                    total_confidence += doc_result.confidence_score
                else:
                    logger.warning(f"Failed to process document: {doc_result.error}")
            
            if not extracted_data:
                return self.create_error_result("Failed to extract data from any documents")
            
            # Calculate average confidence
            avg_confidence = total_confidence / len(extracted_data)
            
            # Consolidate extracted data
            consolidated_data = self._consolidate_data(extracted_data)
            
            self.log_action("document_extraction", {
                "documents_processed": len(documents),
                "successful_extractions": len(extracted_data),
                "average_confidence": avg_confidence
            })
            
            return self.create_success_result(consolidated_data, avg_confidence)
            
        except Exception as e:
            logger.error(f"Extraction agent error: {e}")
            return self.create_error_result(f"Extraction failed: {str(e)}")
    
    async def _process_document(self, document: Dict[str, Any]) -> ProcessingResult:
        """Process individual document based on its type"""
        try:
            file_path = document.get('file_path')
            file_type = document.get('file_type', '').lower()
            
            if not file_path or not Path(file_path).exists():
                return self.create_error_result(f"File not found: {file_path}")
            
            if file_type == 'pdf':
                return await self._extract_from_pdf(file_path)
            elif file_type in ['png', 'jpg', 'jpeg']:
                return await self._extract_from_image(file_path)
            elif file_type in ['xlsx', 'xls']:
                return await self._extract_from_excel(file_path)
            elif file_type == 'txt':
                return await self._extract_from_text(file_path)
            else:
                return self.create_error_result(f"Unsupported file type: {file_type}")
                
        except Exception as e:
            logger.error(f"Document processing error: {e}")
            return self.create_error_result(f"Document processing failed: {str(e)}")
    
    async def _extract_from_pdf(self, file_path: str) -> ProcessingResult:
        """Extract text and structured data from PDF"""
        try:
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                text_content = ""
                
                for page in pdf_reader.pages:
                    text_content += page.extract_text()
                
                # Extract structured information
                structured_data = self._parse_text_for_structured_data(text_content)
                
                # Analyze PDF content for relevance
                pdf_analysis = self._analyze_pdf_content(text_content, len(pdf_reader.pages))
                
                return self.create_success_result({
                    'content_type': 'pdf',
                    'raw_text': text_content,
                    'structured_data': structured_data,
                    'pages': len(pdf_reader.pages),
                    'pdf_analysis': pdf_analysis
                }, confidence=0.9)
                
        except Exception as e:
            logger.error(f"PDF extraction error: {e}")
            return self.create_error_result(f"PDF extraction failed: {str(e)}")
    
    def _analyze_pdf_content(self, text_content: str, page_count: int) -> Dict[str, Any]:
        """Analyze PDF content for relevance and validity"""
        analysis = {
            'document_type': 'unknown',
            'relevance_score': 0,
            'suspicious_flags': [],
            'content_indicators': [],
            'page_analysis': {}
        }
        
        try:
            # Analyze page count
            if page_count < 1:
                analysis['suspicious_flags'].append("PDF has no readable pages")
                analysis['relevance_score'] -= 30
            elif page_count == 1:
                analysis['content_indicators'].append("Single page document")
                analysis['relevance_score'] += 5
            else:
                analysis['content_indicators'].append(f"Multi-page document ({page_count} pages)")
                analysis['relevance_score'] += 10
            
            # Analyze text content
            if text_content:
                text_length = len(text_content.strip())
                if text_length < 50:
                    analysis['suspicious_flags'].append("PDF contains very little text - may be fake or corrupted")
                    analysis['relevance_score'] -= 40
                elif text_length > 5000:
                    analysis['content_indicators'].append("Document contains substantial content")
                    analysis['relevance_score'] += 25
                
                # Check for common document keywords
                document_keywords = {
                    'financial': ['bank', 'statement', 'salary', 'income', 'balance', 'account', 'transaction', 'deposit', 'withdrawal'],
                    'identity': ['id', 'passport', 'license', 'certificate', 'national', 'government', 'official'],
                    'employment': ['employment', 'job', 'work', 'company', 'employer', 'position', 'department'],
                    'residence': ['address', 'residence', 'home', 'property', 'rent', 'lease', 'utility']
                }
                
                found_categories = {}
                for category, keywords in document_keywords.items():
                    found_keywords = [keyword for keyword in keywords if keyword.lower() in text_content.lower()]
                    if found_keywords:
                        found_categories[category] = found_keywords
                        analysis['relevance_score'] += len(found_keywords) * 3
                
                if found_categories:
                    analysis['content_indicators'].extend([f"{cat}: {', '.join(kw)}" for cat, kw in found_categories.items()])
                    # Determine document type based on most common category
                    if 'financial' in found_categories:
                        analysis['document_type'] = 'financial_document'
                    elif 'identity' in found_categories:
                        analysis['document_type'] = 'identity_document'
                    elif 'employment' in found_categories:
                        analysis['document_type'] = 'employment_document'
                    elif 'residence' in found_categories:
                        analysis['document_type'] = 'residence_document'
                else:
                    analysis['suspicious_flags'].append("No relevant document keywords found")
                    analysis['relevance_score'] -= 20
                
                # Check for suspicious patterns
                suspicious_patterns = [
                    (r'test\s+document', 'Test document detected'),
                    (r'sample\s+document', 'Sample document detected'),
                    (r'fake\s+document', 'Fake document detected'),
                    (r'dummy\s+data', 'Dummy data detected'),
                    (r'placeholder', 'Placeholder text detected')
                ]
                
                for pattern, message in suspicious_patterns:
                    if re.search(pattern, text_content, re.IGNORECASE):
                        analysis['suspicious_flags'].append(message)
                        analysis['relevance_score'] -= 25
            else:
                analysis['suspicious_flags'].append("PDF contains no extractable text")
                analysis['relevance_score'] -= 50
            
            # Normalize relevance score to 0-100
            analysis['relevance_score'] = max(0, min(100, analysis['relevance_score'] + 50))
            
        except Exception as e:
            logger.warning(f"PDF analysis error: {e}")
            analysis['relevance_score'] = 50  # Default score
        
        return analysis
    
    def _analyze_text_content(self, text_content: str) -> Dict[str, Any]:
        """Analyze text content for relevance and validity"""
        analysis = {
            'document_type': 'unknown',
            'relevance_score': 0,
            'suspicious_flags': [],
            'content_indicators': []
        }
        
        try:
            # Analyze text content length
            text_length = len(text_content.strip())
            if text_length < 50:
                analysis['suspicious_flags'].append("Text file contains very little content - may be fake")
                analysis['relevance_score'] -= 40
            elif text_length > 1000:
                analysis['content_indicators'].append("Document contains substantial content")
                analysis['relevance_score'] += 25
            elif text_length > 500:
                analysis['content_indicators'].append("Document contains moderate content")
                analysis['relevance_score'] += 15
            else:
                analysis['content_indicators'].append("Document contains basic content")
                analysis['relevance_score'] += 5
            
            # Check for common document keywords
            document_keywords = {
                'financial': ['bank', 'statement', 'salary', 'income', 'balance', 'account', 'transaction', 'deposit', 'withdrawal', 'credit', 'debit', 'amount', 'rupee', '₹'],
                'identity': ['id', 'passport', 'license', 'certificate', 'national', 'government', 'official', 'name', 'address'],
                'employment': ['employment', 'job', 'work', 'company', 'employer', 'position', 'department', 'tcs', 'salary'],
                'residence': ['address', 'residence', 'home', 'property', 'rent', 'lease', 'utility', 'marine drive', 'mumbai']
            }
            
            found_categories = {}
            for category, keywords in document_keywords.items():
                found_keywords = [keyword for keyword in keywords if keyword.lower() in text_content.lower()]
                if found_keywords:
                    found_categories[category] = found_keywords
                    analysis['relevance_score'] += len(found_keywords) * 5
            
            if found_categories:
                analysis['content_indicators'].extend([f"{cat}: {', '.join(kw)}" for cat, kw in found_categories.items()])
                # Determine document type based on most common category
                if 'financial' in found_categories:
                    analysis['document_type'] = 'financial_document'
                elif 'identity' in found_categories:
                    analysis['document_type'] = 'identity_document'
                elif 'employment' in found_categories:
                    analysis['document_type'] = 'employment_document'
                elif 'residence' in found_categories:
                    analysis['document_type'] = 'residence_document'
            else:
                analysis['suspicious_flags'].append("No relevant document keywords found")
                analysis['relevance_score'] -= 20
            
            # Check for suspicious patterns
            suspicious_patterns = [
                (r'test\s+document', 'Test document detected'),
                (r'sample\s+document', 'Sample document detected'),
                (r'fake\s+document', 'Fake document detected'),
                (r'dummy\s+data', 'Dummy data detected'),
                (r'placeholder', 'Placeholder text detected')
            ]
            
            for pattern, message in suspicious_patterns:
                if re.search(pattern, text_content, re.IGNORECASE):
                    analysis['suspicious_flags'].append(message)
                    analysis['relevance_score'] -= 25
            
            # Normalize relevance score to 0-100
            analysis['relevance_score'] = max(0, min(100, analysis['relevance_score'] + 50))
            
        except Exception as e:
            logger.warning(f"Text analysis error: {e}")
            analysis['relevance_score'] = 50  # Default score
        
        return analysis
    
    async def _extract_from_image(self, file_path: str) -> ProcessingResult:
        """Extract text from image using OCR"""
        try:
            image = Image.open(file_path)
            
            # Check if tesseract is available
            try:
                text_content = pytesseract.image_to_string(image)
                ocr_available = True
            except Exception as ocr_error:
                logger.warning(f"OCR not available, using basic image analysis: {ocr_error}")
                text_content = ""
                ocr_available = False
            
            # Basic image analysis even without OCR
            image_analysis = self._analyze_image_content(image, text_content, ocr_available)
            
            # Extract structured information
            structured_data = self._parse_text_for_structured_data(text_content)
            
            return self.create_success_result({
                'content_type': 'image',
                'raw_text': text_content,
                'structured_data': structured_data,
                'image_size': image.size,
                'ocr_available': ocr_available,
                'image_analysis': image_analysis
            }, confidence=0.8 if ocr_available else 0.6)
            
        except Exception as e:
            logger.error(f"Image processing error: {e}")
            return self.create_error_result(f"Image processing failed: {str(e)}")
    
    def _analyze_image_content(self, image: Image.Image, text_content: str, ocr_available: bool) -> Dict[str, Any]:
        """Analyze image content for relevance and validity"""
        analysis = {
            'image_type': 'unknown',
            'relevance_score': 0,
            'suspicious_flags': [],
            'content_indicators': []
        }
        
        try:
            # Analyze image dimensions
            width, height = image.size
            aspect_ratio = width / height if height > 0 else 0
            
            # Check for suspicious image characteristics
            if width < 100 or height < 100:
                analysis['suspicious_flags'].append("Image too small - may be fake")
                analysis['relevance_score'] -= 20
            
            if aspect_ratio < 0.5 or aspect_ratio > 3:
                analysis['suspicious_flags'].append("Unusual aspect ratio - may be manipulated")
                analysis['relevance_score'] -= 15
            
            # Analyze image mode and colors
            if image.mode == 'RGB':
                # Convert to analyze color distribution
                colors = image.getcolors(maxcolors=1000)
                if colors:
                    total_pixels = sum(count for count, _ in colors)
                    if total_pixels > 0:
                        # Check for suspicious color patterns
                        black_pixels = sum(count for count, color in colors if sum(color) < 30)
                        white_pixels = sum(count for count, color in colors if sum(color) > 700)
                        
                        if black_pixels / total_pixels > 0.8:
                            analysis['suspicious_flags'].append("Image too dark - may be corrupted")
                            analysis['relevance_score'] -= 25
                        
                        if white_pixels / total_pixels > 0.8:
                            analysis['suspicious_flags'].append("Image too bright - may be blank")
                            analysis['relevance_score'] -= 25
            
            # Analyze text content if OCR was available
            if ocr_available and text_content:
                text_length = len(text_content.strip())
                if text_length < 10:
                    analysis['suspicious_flags'].append("Very little text extracted - may be fake document")
                    analysis['relevance_score'] -= 30
                elif text_length > 1000:
                    analysis['content_indicators'].append("Document contains substantial text content")
                    analysis['relevance_score'] += 20
                
                # Check for common document keywords
                document_keywords = ['bank', 'statement', 'salary', 'income', 'id', 'passport', 'license', 'certificate', 'bill', 'receipt']
                found_keywords = [keyword for keyword in document_keywords if keyword.lower() in text_content.lower()]
                if found_keywords:
                    analysis['content_indicators'].extend(found_keywords)
                    analysis['relevance_score'] += len(found_keywords) * 5
                    analysis['image_type'] = 'document'
                else:
                    analysis['suspicious_flags'].append("No relevant document keywords found")
                    analysis['relevance_score'] -= 15
            else:
                # No OCR available - use basic image analysis
                analysis['relevance_score'] = max(0, analysis['relevance_score'])
                analysis['content_indicators'].append("OCR not available - limited content analysis")
            
            # Normalize relevance score to 0-100
            analysis['relevance_score'] = max(0, min(100, analysis['relevance_score'] + 50))
            
        except Exception as e:
            logger.warning(f"Image analysis error: {e}")
            analysis['relevance_score'] = 50  # Default score
        
        return analysis
    
    async def _extract_from_excel(self, file_path: str) -> ProcessingResult:
        """Extract data from Excel files"""
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            sheet = workbook.active
            
            data = []
            for row in sheet.iter_rows(values_only=True):
                if any(cell is not None for cell in row):
                    data.append(row)
            
            # Try to identify headers and structure
            structured_data = self._structure_excel_data(data)
            
            return self.create_success_result({
                'content_type': 'excel',
                'raw_data': data,
                'structured_data': structured_data,
                'sheets': len(workbook.sheetnames)
            }, confidence=0.95)
            
        except Exception as e:
            logger.error(f"Excel extraction error: {e}")
            return self.create_error_result(f"Excel extraction failed: {str(e)}")
    
    async def _extract_from_text(self, file_path: str) -> ProcessingResult:
        """Extract data from text files"""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                text_content = file.read()
            
            structured_data = self._parse_text_for_structured_data(text_content)
            
            # Analyze text content for relevance
            text_analysis = self._analyze_text_content(text_content)
            
            return self.create_success_result({
                'content_type': 'text',
                'raw_text': text_content,
                'structured_data': structured_data,
                'text_analysis': text_analysis
            }, confidence=0.9)
            
        except Exception as e:
            logger.error(f"Text extraction error: {e}")
            return self.create_error_result(f"Text extraction failed: {str(e)}")
    
    def _parse_text_for_structured_data(self, text: str) -> Dict[str, Any]:
        """Parse text to extract structured information"""
        structured_data = {}
        
        # Extract email addresses
        email_pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
        emails = re.findall(email_pattern, text)
        if emails:
            structured_data['emails'] = emails
        
        # Extract phone numbers
        phone_pattern = r'(\+\d{1,3}[-.\s]?)?\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}'
        phones = re.findall(phone_pattern, text)
        if phones:
            structured_data['phone_numbers'] = phones
        
        # Extract monetary amounts
        money_pattern = r'\$[\d,]+\.?\d*'
        money_amounts = re.findall(money_pattern, text)
        if money_amounts:
            structured_data['monetary_amounts'] = money_amounts
        
        # Extract dates
        date_pattern = r'\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\b'
        dates = re.findall(date_pattern, text)
        if dates:
            structured_data['dates'] = dates
        
        return structured_data
    
    def _structure_excel_data(self, data: List[tuple]) -> Dict[str, Any]:
        """Structure Excel data into meaningful format"""
        if not data:
            return {}
        
        # Assume first row contains headers
        headers = data[0] if data else []
        rows = data[1:] if len(data) > 1 else []
        
        structured_data = {
            'headers': headers,
            'rows': rows,
            'column_count': len(headers) if headers else 0,
            'row_count': len(rows)
        }
        
        # Try to identify key columns
        for i, header in enumerate(headers):
            if header:
                header_lower = str(header).lower()
                if 'name' in header_lower:
                    structured_data['name_column'] = i
                elif 'income' in header_lower or 'salary' in header_lower:
                    structured_data['income_column'] = i
                elif 'date' in header_lower:
                    structured_data['date_column'] = i
        
        return structured_data
    
    def _consolidate_data(self, extracted_data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Consolidate data from multiple documents"""
        consolidated = {
            'total_documents': len(extracted_data),
            'document_types': [],
            'all_text_content': [],
            'structured_data': {},
            'confidence_scores': []
        }
        
        for data in extracted_data:
            consolidated['document_types'].append(data.get('content_type', 'unknown'))
            consolidated['all_text_content'].append(data.get('raw_text', ''))
            consolidated['confidence_scores'].append(data.get('confidence', 0.0))
            
            # Merge structured data
            if 'structured_data' in data:
                for key, value in data['structured_data'].items():
                    if key not in consolidated['structured_data']:
                        consolidated['structured_data'][key] = []
                    if isinstance(value, list):
                        consolidated['structured_data'][key].extend(value)
                    else:
                        consolidated['structured_data'][key].append(value)
        
        return consolidated 